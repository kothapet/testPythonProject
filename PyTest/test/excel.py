'''
Created on Nov 2, 2017

@author: anand
'''

# python -m pip install -U pip setuptools

if __name__ == '__main__':
    pass


import openpyxl
import jaydebeapi


wb = openpyxl.load_workbook('D:\\Users\\anand\\Documents\\orders.xlsx')

ws = wb.get_sheet_names()

print(ws)

print(ws[0])

sheet = wb.get_sheet_by_name(ws[0])

print(sheet)

maxrow, maxcolumn  = sheet.max_row, sheet.max_column

for i in range(1, maxrow):
    for j in range(1, maxcolumn):
        print(i, j, sheet.cell(row=i, column=j).value)



conn = jaydebeapi.connect("org.hsqldb.jdbcDriver",
                          "jdbc:hsqldb:mem:.",
                          ["SA", ""],
                          "/path/to/hsqldb.jar",)

curs = conn.cursor()
curs.execute('create table CUSTOMER'
               '("CUST_ID" INTEGER not null,'
               ' "NAME" VARCHAR not null,'
               ' primary key ("CUST_ID"))'
            )
curs.execute("insert into CUSTOMER values (1, 'John')")
curs.execute("select * from CUSTOMER")
curs.fetchall()
curs.close()
conn.close()

